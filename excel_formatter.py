import settings
from models import Person, Nickname
from typing import Any, List, Tuple, Union
from openpyxl.styles import Font

import xlrd

from openpyxl import Workbook

FileName = str
Range = Tuple[int, int]
font = Font("Times new Roman", 9)
semestr_index = 4
course_index = 2
name_index = 21
max_rows_in_individual_load = 55
lection_index = 8
Error = Any
IsError = bool


class Teacher:
    def __init__(self, name, nicknames):
        self.name = name
        self.nicknames = [name] + nicknames

        self.all = []


    def add(self, element):
        del element[21]
        self.all.append(element)

headers_rows = [
    "N  п/п",
    "Ф.И.О. преподавателя",
    "Должность",
    "Ставка",
    "Семестр",
    "Лекции",
    "Практ. сем. занятия",
    "Лабор. занятия",
    "Контр. работы",
    "Консуль-    тации",
    "Экзамены",
    "Курс.        работы/ НИР",
    "Практика",
    "Рецензирование (маг.,спец.), экспертиза",
    "Рук-во ВКР,  НКР",
    "ГЭК",
    "Рук-во ОПОП",
    "Всего",
]

create_result_load_io = "FullLoad.xls"
create_first_load_io = "FirstLoad.xls"
create_second_load_io = "SecondLoad.xls"

all_results_summ = []

all_sums = {
    # 'name': [first_sum, second_sum]
}

def add_style(ws):
    for s in "ABCDEFGHIJKLMNOPQRSTU":
        for i in range(1, max_rows_in_individual_load):
            el = ws[f"{s}{i}"]
            el.font = font

def is_digit(obj):
    if isinstance(obj, int) or isinstance(obj, float):
        return True

    if obj.isdigit():
       return True
    else:
        try:
            float(obj)
            return True
        except ValueError:
            return False

def list_sum(a, b):
    if len(a) != len(b):
        # print(len(a), len(b))
        raise ValueError()

    new = [[0] for _ in range(len(a))]
    for i in range(len(a)):
        new_s = 0
        if is_digit(a[i]):
            new_s += a[i]

    
        if is_digit(b[i]):
            new_s += b[i]
    
        
        new[i] = new_s
    
    return new

def get_all_nicknames():
    read_io = settings.READ_FILE
    headers = settings.HEADERS

    data = get_sheets(read_io, headers)
    nicks = []
    for el in data:

        if el[name_index] not in ["", " "] + nicks and not el[name_index].startswith("Unnamed"):
            nicks.append(el[name_index])

    return nicks

def init_teachers(teachers:dict):
    pers = []
    
    for element in teachers.items():
        pers.append(Teacher(name = element[0], nicknames = element[1]))

    return pers

def get_sheets(read_io, headers):
    rb = xlrd.open_workbook(read_io, formatting_info = True)
    sheet = rb.sheet_by_index(0)
    data = []
    for i in range(*headers):
        try:
            data.append(sheet.row_values(i))
        except:
            # print("Последняя колонка номер", i)
            break
    return data

def update_teachers(data, teachers) -> tuple[Union[list[Teacher], List[str]], IsError]:
    # соотносит нагрузку с учителем
    not_found = []
    for element in data:
        for teacher in teachers:
            if element[name_index] in teacher.nicknames:
                teacher.add(element)
                break

        else:
            if element[21] not in not_found and element[21]:
                not_found.append(element[21])

    if not not_found:
        return (teachers, 0)

    return (not_found, 1)

def create_inividual_load(teachers, write_io):
    wb = Workbook()
    all_results = []
    global all_results_summ
    all_results_summ = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    for teacher in teachers:
        ws = wb.create_sheet(teacher.name)
        ws.title = teacher.name

        data = [
            ["  " * 100 +  "Сведения о планируемой учебной нагрузке"],
            [" ", "Институт", "ИЕН", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", "Должность" ],
            [" ", "Кафедра", "Экология и природопользования", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", "Ставка"],
            [" ", "Фамилия", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", "Ученая степень"],
            [" ", "Имя", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", "Штат./Совм.,работод."],
            [" ", "Отчество", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", "Учебный год"], 
            ["Дисциплина/вид работы", "Институт", "Курс", "Форма обучения", "Семестр", "Кол-во недель",	"Кол-во\nстудентов", "Кол-во\nгрупп", "Лекции", "Практ. сем. занятия", "Лабор. занятия", "Контр. работы", "Консуль-тации", "Экзамены", "Курс.работы/ \nНИР", "Практика", "Рецензирование (маг.,спец.), экспертиза",	"Рук-во ВКР, НКР",  "ГЭК", "Рук-во ОПОП",	"Всего"]
        ]

        first_sum = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        second_sum = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

        #first semestr
        data.append(["  " * 100 +  "Бакалавриат/Специалитет"])
        data.append([])
        for element in teacher.all:

            if "ОПОП" in element[0] and 'б' in str(element[0]):
                data.append(element)
                first_sum = list_sum(first_sum, element[8:21])

                continue

            if ("б" in str(element[course_index]) or "БЖД" in element[0]) and int(element[semestr_index]) % 2 == 1:
                data.append(element)
                first_sum = list_sum(first_sum, element[8:21])

        data.append(["  " * 100 +  "Магистратура"])
        data.append([])
        for element in teacher.all:
            if "ОПОП" in element[0] and 'м' in str(element[0]):
                data.append(element)
                first_sum = list_sum(first_sum, element[8:21])
                continue
            if "м" in str(element[course_index]) and int(element[semestr_index]) % 2 == 1:
                data.append(element)
                first_sum = list_sum(first_sum, element[8:21])

        data.append(["  " * 100 +  "Аспирантура"])
        data.append([])
        for element in teacher.all:
            if "ОПОП" in element[0]:
                continue
            try:
                if "м" not in str(element[course_index]) and "б" not in str(element[course_index]) and "БЖД" not in element[0] and int(element[semestr_index]) % 2 == 1:
                    data.append(element)
                    first_sum = list_sum(first_sum, element[8:21])
            except ValueError:
                if "бакалавр" in element[0] or "магистр" in element[0]:
                    continue

                if "аспирант" in element[0]:
                    data.append(element)
                    first_sum = list_sum(first_sum, element[8:21])

        res = ["", "", "", "", "", "", "", "Всего за 1 семестр"] + first_sum
        data.append(res)
        data.append([])

        #second semestr
        data.append([])
        data.append(["  " * 100 +  "Бакалавриат/Специалитет"])
        data.append([])
        for element in teacher.all:
            if "ОПОП" in element[0]:
                continue
            if ("б" in str(element[course_index]) or "БЖД" in element[0]) and int(element[semestr_index]) % 2 == 0:
                data.append(element)
                second_sum = list_sum(second_sum, element[8:21])

        data.append(["  " * 100 +  "Магистратура"])
        data.append([])
        for element in teacher.all:
            if "ОПОП" in element[0]:
                continue
            if "м" in str(element[course_index]) and int(element[semestr_index]) % 2 == 0:
                data.append(element)
                second_sum = list_sum(second_sum, element[8:21])

        data.append(["  " * 100 +  "Аспирантура"])
        data.append([])
        for element in teacher.all:
            if "ОПОП" in element[0]:
                continue

            try:
                if "м" not in str(element[course_index]) and "б" not in str(element[course_index]) and "БЖД" not in str(element[0]) and int(element[semestr_index]) % 2 == 0:
                    data.append(element)
                    second_sum = list_sum(second_sum, element[8:21])
            except:
                pass

        res = ["", "", "", "", "", "", "", "Всего за 2 семестр"] + second_sum
        data.append(res)
        data.append([])

        full_load = list_sum(second_sum, first_sum)
        res = ["", "", "", "", "", "", "", "Итого за год"] + full_load
        data.append(res)
        data.append([])

        global all_sums
        all_sums[teacher.name] = [first_sum, second_sum]
        all_results_summ = list_sum(all_results_summ, full_load)
        all_results.append(full_load + [teacher.name])

        #full result
        for row in data:
            ws.append(row)

        add_style(ws)

    ws = wb.create_sheet("Итог")
    ws.title = "Итог"

    data = all_results + [all_results_summ + ["Всего"]]
    for row in data:
        ws.append(row)

    add_style(ws)
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(write_io)

def create_result_load(write_io):
    wb = Workbook()
    ws = wb.create_sheet("Распред. уч. Нагр. Кафедры")
    ws.title = "Распред. уч. Нагр. Кафедры"

    data = [headers_rows.copy()]
    for people, value in all_sums.items():
        for key in range(2):
            val = value[key]

            data.append(
                [
                    " ",
                    people,
                    " ",
                    " ",
                    str(key + 1),
                    *[el for el in val]
                ]
            )
    data.append(
        [
            ' ',
            ' ',
            'Итого',
            ' ',
            ' ',
            *all_results_summ
        ]
    )

    for row in data:
        ws.append(row)

    add_style(ws)
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(write_io)

def create_part_result_load(write_io, key:int):
    wb = Workbook()
    ws = wb.create_sheet("Отчет")
    ws.title = "Отчет"

    data = [[*headers_rows.copy(), "+/-", "Причины изменения,  № протокола зас. каф.", "Подпись"]]
    del data[0][4]
    # print(data)

    all_summ = [0,0,0,0,0,0,0,0,0,0,0,0,0]

    for people, value in all_sums.items():

        val = value[key]
        for index, v in enumerate(val):
            all_summ[index] += v

        data.append(
            [
                " ",
                people,
                " ",
                " ",
                *[el for el in val]
            ]
        )
    data.append(
        [
            ' ',
            ' ',
            'Итого',
            ' ',
            *all_summ,
        ]
    )

    #full result
    for row in data:
        ws.append(row)

    add_style(ws)
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    wb.save(write_io)

def get_teachers() -> dict[str, list[str]]:
    teachers_nicknames = {}

    for teacher in Person.select():
        nicks = [
            nick.name
            for nick in Nickname.select().where(Nickname.person==teacher.id)
        ]
        nicks.append(teacher.name)
        nicks = list(set(nicks))

        teachers_nicknames[teacher.name] = nicks

    return teachers_nicknames


def main():
    write_io = settings.WRITE_FILE
    read_io = settings.READ_FILE
    headers = settings.HEADERS

    teachers = get_teachers()

    data = get_sheets(read_io, headers)
    teachers = init_teachers(teachers)

    teachers, error = update_teachers(data, teachers)
    if error:
        # if error not found nicknames
        return "Обнаружены неизвестные псевдонимы:\n" + '\n'.join('"'+t+'"' for t in teachers)

    try:
        create_inividual_load(teachers, write_io)
        create_result_load(create_result_load_io)
        create_part_result_load("osen.xls", 0)
        create_part_result_load("vesna.xls", 1)
    except PermissionError:
        return "Ошибка: возможно вы не закрыли редактируемый файл."

    return "Успешно"
